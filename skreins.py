#!/usr/bin/env python3

import glob
import os
import pandas 
import subprocess
import sys
import openpyxl
import csv
from itertools import chain
#import xlsxwriter

import clean_skript_V3


#convert lb and clean column to .txt
excel_files = glob.glob(os.getcwd() + '/*.xlsx') 

for excel in excel_files:
    base_name = excel.split('.')[0]
    convex = base_name + '_CONVED.txt'
#apply clean-skript_V3.py
    df = pandas.read_excel(excel)
    df[['lb','dipl']].to_csv(convex, index=False, sep='\t') 
    output = base_name + '_CLEANED.csv'
    #subprocess.call(['python3', 'clean-skript_V3.py', convex, output], shell=True
    clean_skript_V3.main(["", convex, output])
#rename cleaned 'dipl' column as string 
    cleaned = base_name + '_CLEANDIPLED.csv'
    col = pandas.read_csv(output, error_bad_lines=False, sep='\t')
    col_rename = col.rename(columns={'dipl':'clean'}, index = None)
    cleaned_col = col_rename[['clean']].to_csv(cleaned, index = None, header=True)
    origclean = pandas.read_csv(cleaned, error_bad_lines=False, sep='\t')
    finalclean = base_name + '_FINAL.xlsx'
    cleaned_to_excel = origclean.to_excel(finalclean, index = None)
    read_final_clean = pandas.read_excel(finalclean)
#insert new 'clean' column (delete old one)
    wb = openpyxl.load_workbook(filename=excel, read_only=False)
    ws = wb.active
    dipl_col_index = None
    clean_col_index = None
    
    for e, col in enumerate(ws.iter_cols()):     
        if col[0].value == 'clean':
           clean_col_index = e+1
    if clean_col_index is not None:
        ws.delete_cols(clean_col_index)
    for e, col in enumerate(ws.iter_cols()):
        if col[0].value == 'dipl':
            dipl_col_index = e+1
    ws.insert_cols(dipl_col_index+1)

#need to save and reopen,
#otherwise new inserted column won't be found when deleting merged cells
    wb.save("TEMP.xlsx")   
    wb = openpyxl.load_workbook(filename="TEMP.xlsx", read_only=False)
    ws = wb.active
    
#find and unmerge merged cells
    merged_cells = list(ws.merged_cells.ranges)
    for cr in merged_cells:    
        if(cr.min_col==dipl_col_index+1): 
            ws.unmerge_cells(range_string=str(cr))

    
#bring edited 'clean' column back to origine
    with open(cleaned) as f: 
        reader = csv.reader(f, delimiter=';')
        for i, row in enumerate(reader):
                c = ws.cell(row=i+1, column=dipl_col_index+1)
                c.value = row[0]

    last_value = -1
    merging = False
    for row in ws.iter_rows(min_col=dipl_col_index+1, max_col=dipl_col_index+1): 
        cell = row[0]
        #print("row ", row.index, "  value ", cell.value)
        if not cell.value:
            if not merging:
                merging = True
                last_value = cell.row-1

        elif merging:
            ws.merge_cells(start_row=last_value, start_column=dipl_col_index+1, end_row=cell.row-1, end_column=dipl_col_index+1)
            merging = False


    wb.save(excel)


 
