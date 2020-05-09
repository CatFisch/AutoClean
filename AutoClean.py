#!/usr/bin/env python3

#Copyright 2020 Catharina Fischer and Alexandra Tichauer

#Licensed under the Apache License, Version 2.0 (the "License");
#you may not use this file except in compliance with the License.
#You may obtain a copy of the License at

    #http://www.apache.org/licenses/LICENSE-2.0

#Unless required by applicable law or agreed to in writing, software
#distributed under the License is distributed on an "AS IS" BASIS,
#WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#See the License for the specific language governing permissions and
#limitations under the License.



import glob
import os
import pandas 
import subprocess
import sys
import openpyxl
import csv
from itertools import chain
import shutil
import clean_skript_V3
import argparse

#convert .xlsx files in .txt  (start function for optionality of input files)          
def process_file(excel):
    base_name = excel.split('.')[0]
    convex = base_name + '_CONVED.txt'

#apply clean-skript_V3.py
    df = pandas.read_excel(excel)
    df[['lb','dipl']].to_csv(convex, index=False, sep='\t') 
    output = base_name + '_CLEANED.csv'
    clean_skript_V3.main(['', convex, output])
    
#rename cleaned 'dipl' column as string 
    cleaned = base_name + '_CLEANDIPLED.csv'
    col = pandas.read_csv(output, error_bad_lines=False, sep='\t')
    col_rename = col.rename(columns={'dipl':'clean'}, index = None)
    cleaned_col = col_rename[['clean']].to_csv(cleaned, index = None, header=True)
    origclean = pandas.read_csv(cleaned, error_bad_lines=False, sep='\t')
    finalclean = base_name + '_CLEANED.xlsx'
    cleaned_to_excel = origclean.to_excel(finalclean, index = None)
    read_final_clean = pandas.read_excel(finalclean)
    
#insert new 'clean' column (delete old one)
    wb = openpyxl.load_workbook(filename=excel, read_only=False)
    ws = wb.worksheets[0]
    dipl_col_index = None
    clean_col_index = None
    
    for e, col in enumerate(ws.iter_cols()):     
        if col[0].value == 'clean':
           clean_col_index = e+1
    if clean_col_index is not None:
        ws.delete_cols(clean_col_index)
    for e, col in enumerate(ws.iter_cols()):
        if col[0].value == 'dipl':
            dipl_col_index = e+1
    ws.insert_cols(dipl_col_index+1)

#need to save and reopen (otherwise new inserted column won't be found when deleting merged cells)
    wb.save(base_name + 'TEMP.xlsx')   
    wb = openpyxl.load_workbook(filename= base_name + 'TEMP.xlsx', read_only=False)
    ws = wb.worksheets[0] 
    
#find and unmerge merged cells
    merged_cells = list(ws.merged_cells.ranges)
    for cr in merged_cells:    
        if(cr.min_col==dipl_col_index+1): 
            ws.unmerge_cells(range_string=str(cr))

#bring edited 'clean' column back to origine
    with open(cleaned) as f: 
        reader = csv.reader(f, delimiter=';')
        for i, row in enumerate(reader):
                c = ws.cell(row=i+1, column=dipl_col_index+1)
                c.value = row[0]

    last_value = -1
    merging = False
    for row in ws.iter_rows(min_col=dipl_col_index+1, max_col=dipl_col_index+1): 
        cell = row[0]
        if not cell.value:
            if not merging:
                merging = True
                last_value = cell.row-1

        elif merging:
            ws.merge_cells(start_row=last_value,
                           start_column=dipl_col_index+1,
                           end_row=cell.row-1,
                           end_column=dipl_col_index+1)
            merging = False

    wb.save(excel)
    
#move output to 'Collected_Output'
    try:
        os.mkdir('Collected_Output')
    except OSError:
        pass
    
    dest = 'Collected_Output'
    files = os.listdir(os.getcwd())

    for f in files:
        if (f.startswith("__pycache__") or
            f.endswith(".csv") or
            f.endswith(".txt") or
            f.endswith("TEMP.xlsx") or
            f.endswith("_CLEANED.xlsx")):
            shutil.move(os.path.join(os.getcwd(), f), os.path.join(dest, f))
    #print final message
    print(os.path.basename(base_name), "was cleaned successfully")

#end function for optionality of input files
parser = argparse.ArgumentParser()
parser.add_argument("--table", nargs='*')
args = parser.parse_args()

#clean list of files in folder
if args.table:
    for f in args.table:
        process_file(f)
#clean all files in folder        
else:
    excel_files = glob.glob(os.getcwd() + '/*.xlsx')
    
    for excel in excel_files:
        process_file(excel)

print("\n selected files are cleand altogeter \n ฅ^•ﻌ•^ฅ")
  
